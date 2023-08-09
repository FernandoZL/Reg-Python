import os
import datetime
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import gspread
from tabulate import tabulate
import platform

# Configuración de las credenciales
CLIENT_ID = "1089617708991-6guk84qua1u2gv8v1poohj60najt97hk.apps.googleusercontent.com"
CLIENT_SECRET = "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"
REDIRECT_URI = "http://localhost"
SCOPE = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/spreadsheets",
         "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]

def obtener_credenciales():
    from oauth2client import tools
    from oauth2client.file import Storage
    from oauth2client.client import OAuth2WebServerFlow

    flow = OAuth2WebServerFlow(client_id=CLIENT_ID, client_secret=CLIENT_SECRET, scope=SCOPE, redirect_uri=REDIRECT_URI)
    storage = Storage('credentials.dat')
    credentials = storage.get()
    if credentials is None or credentials.invalid:
        credentials = tools.run_flow(flow, storage)

    return credentials

def obtener_numero_turno():
    archivo = "RegistrosTickets.xlsx"
    numero_turno = ""

    if os.path.exists(archivo):
        workbook = load_workbook(filename=archivo)
        worksheet = workbook.active
        if worksheet.max_row > 1:
            ultimo_numero_turno = int(worksheet.cell(row=worksheet.max_row, column=1).value)
            numero_turno = str(ultimo_numero_turno + 1).zfill(6)
        else:
            numero_turno = "000001"
    else:
        numero_turno = "000001"

    return numero_turno

def formatear_numero_ingreso(numero_turno):
    return "Turno: " + numero_turno.zfill(6)

def formatear_fecha_hora(fecha_hora_turno):
    fecha_hora = datetime.datetime.strptime(fecha_hora_turno, "%Y-%m-%d %I:%M %p")
    return "Fecha y Hora del turno:\n" + fecha_hora.strftime("%Y-%m-%d %I:%M %p")

def imprimir_ticket(numero_turno, nombre_piloto, no_lic, placas, empresa, origen, fecha_hora_turno):
    carpeta_salida = "Turnos Registrados"
    if not os.path.exists(carpeta_salida):
        os.mkdir(carpeta_salida)

    archivo_salida = os.path.join(carpeta_salida, f"Turno_{numero_turno}.pdf")  # Nombre del archivo con el número de turno
    c = canvas.Canvas(archivo_salida, pagesize=letter)
    c.setFont("Helvetica", 12)

    y = 700

    c.drawString(100, y, "=== BIENVENIDO ===")
    y -= 20
    c.setFont("Helvetica-Bold", 14)
    c.drawString(100, y, formatear_numero_ingreso(numero_turno))

    # Resto del código para dibujar el ticket...

    c.save()
    print(f"Turno impreso y guardado como {archivo_salida}.")

def guardar_registro_en_excel(numero_turno, nombre_piloto, no_lic, placas, empresa, origen, fecha_hora_turno):
    archivo = "RegistrosTickets.xlsx"
    if os.path.exists(archivo):
        workbook = load_workbook(filename=archivo)
    else:
        workbook = Workbook()

    worksheet = workbook.active
    if worksheet.title != "Registros":
        worksheet.title = "Registros"
        worksheet.append(["Turno", "Nombre de piloto", "No. Licencia", "Placas", "Empresa", "Origen", "Fecha y Hora del turno"])

    fila = (numero_turno, nombre_piloto, no_lic, placas, empresa, origen, fecha_hora_turno)
    worksheet.append(fila)

    workbook.save(filename=archivo)
    print("Registro guardado en el archivo Excel.")

def guardar_registro_en_google_sheets(numero_turno, nombre_piloto, no_lic, placas, empresa, origen, fecha_hora_turno):
    try:
        creds = obtener_credenciales()
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_url("https://docs.google.com/spreadsheets/d/1wqVsSCL0ccH3aGNVd1XWhVOFIUwDgfqLS-toJVWBZ4o")
        worksheet = spreadsheet.worksheet("Registros")  # Asegúrate de que la hoja se llame "Registros"

        fila = [numero_turno, nombre_piloto, no_lic, placas, empresa, origen, fecha_hora_turno]
        worksheet.append_row(fila)

        print("Registro guardado en Google Sheets.")
    except Exception as e:
        print("Error al guardar en Google Sheets:", e)


def main():
    limpiar_pantalla()  # Limpia la pantalla antes de mostrar los registros
    
    print("=== Sistema de Turnos ===")
    
    while True:
        opcion = mostrar_menu()
        
        if opcion == "1":
            ver_registros()
        elif opcion == "2":
            ingresar_nuevo_registro()
        elif opcion == "3":
            generar_reporte()
        elif opcion == "4":
            break
        else:
            print("Opcion invalida. Por favor, selecciona una opcion valida.")


def mostrar_menu():
    limpiar_pantalla()  # Limpia la pantalla antes de mostrar los registros
    
    print("\n=== Menu ===")
    print("1. Ver registros")
    print("2. Ingresar nuevo registro")
    print("3. Imprimir reporte")
    print("4. Salir")
    opcion = input("Selecciona una opcion: ")
    return opcion


def limpiar_pantalla():
    os.system('cls' if os.name == 'nt' else 'clear')


def ver_registros():
    limpiar_pantalla()  # Limpia la pantalla antes de mostrar los registros
    
    archivo = "RegistrosTickets.xlsx"
    if os.path.exists(archivo):
        workbook = load_workbook(filename=archivo)
        worksheet = workbook.active

        registros = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            registros.append(row)

        if registros:
            headers = ["Turno", "Nombre de piloto", "NO. LICENCIA", "Placas", "Empresa", "Origen", "Fecha y Hora del turno"]
            tabla = tabulate(registros, headers, tablefmt="grid")
            print("\n=== Registros ===")
            print(tabla)
        else:
            print("No hay registros para mostrar.")

        input("\nPresiona Enter para regresar al menu principal...")
        limpiar_pantalla()  # Limpia la pantalla antes de mostrar los registros

def generar_reporte():
    limpiar_pantalla()  # Limpia la pantalla antes de mostrar los registros
    
    archivo = "RegistrosTickets.xlsx"
    if os.path.exists(archivo):
        workbook = load_workbook(filename=archivo)
        worksheet = workbook.active

        registros = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            registros.append(row)

        if registros:
            headers = ["Turno", "Nombre de piloto", "NO. LICENCIA", "Placas", "Empresa", "Origen", "Fecha y Hora del turno"]
            tabla = tabulate(registros, headers, tablefmt="grid")
            print("\n=== Reporte de Registros ===")
            print(tabla)

            pdf_reporte = "Reporte_Registros.pdf"
            c = canvas.Canvas(pdf_reporte, pagesize=letter)
            c.setFont("Helvetica", 12)

            y = 750
            for row in registros:
                c.drawString(100, y, "Turno: " + str(row[0]))
                y -= 20
                c.setFont("Helvetica-Bold", 12)
                c.drawString(100, y, "Nombre de piloto: " + str(row[1]))
                y -= 15
                c.drawString(100, y, "NO. LICENCIA: " + str(row[2]))
                y -= 15
                c.drawString(100, y, "Placas: " + str(row[3]))
                y -= 15
                c.drawString(100, y, "Empresa: " + str(row[4]))
                y -= 15
                c.drawString(100, y, "Origen: " + str(row[5]))
                y -= 15
                c.drawString(100, y, "Fecha y Hora del turno: " + str(row[6]))
                y -= 50

            c.save()
            print(f"Reporte generado y guardado como {pdf_reporte}.")
        else:
            print("No hay registros para generar el reporte.")

        input("\nPresiona Enter para regresar al menu principal...")
        limpiar_pantalla()  # Limpia la pantalla antes de mostrar los registros



def ingresar_nuevo_registro():
    limpiar_pantalla()  # Limpia la pantalla antes de mostrar los registros
    
    print("\n=== Ingresar Nuevo Registro ===")
    
    nombre_piloto = input("Nombre de piloto: ")
    no_lic = input("No. Licencia: ")
    placas = input("Placas: ")
    empresa = input("Empresa: ")
    origen = input("Origen (Local/Importado): ")

    fecha_hora_actual = datetime.datetime.now()
    fecha_hora_turno = fecha_hora_actual.strftime("%Y-%m-%d %I:%M %p")

    numero_turno = obtener_numero_turno()

    print("\n=== Datos Ingresados ===")
    print("Turno:", formatear_numero_ingreso(numero_turno))
    print("Nombre de piloto:", nombre_piloto)
    print("NO. LICENCIA:", no_lic)
    print("Placas:", placas)
    print("Empresa:", empresa)
    print("Origen:", origen)
    print("Fecha y Hora del turno:", formatear_fecha_hora(fecha_hora_turno))

    imprimir_ticket(numero_turno, nombre_piloto, no_lic, placas, empresa, origen, fecha_hora_turno)
    guardar_registro_en_excel(numero_turno, nombre_piloto, no_lic, placas, empresa, origen, fecha_hora_turno)
    guardar_registro_en_google_sheets(numero_turno, nombre_piloto, no_lic, placas, empresa, origen, fecha_hora_turno)

    limpiar_pantalla()  # Limpia la pantalla antes de mostrar los registros


def main():
    limpiar_pantalla()  # Limpia la pantalla antes de mostrar los registros
    
    print("=== Sistema de Turnos ===")
    
    while True:
        opcion = mostrar_menu()
        
        if opcion == "1":
            ver_registros()
        elif opcion == "2":
            ingresar_nuevo_registro()
        elif opcion == "3":
            generar_reporte()
        elif opcion == "4":
            break
        else:
            print("Opcion invalida. Por favor, selecciona una opcion valida.")
            limpiar_pantalla()  # Limpia la pantalla antes de mostrar los registros

if __name__ == "__main__":
    main()
